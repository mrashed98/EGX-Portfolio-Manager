from io import BytesIO
from datetime import datetime
from typing import List, Optional, Dict, Any
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.portfolio import Portfolio
from app.models.strategy import Strategy
from app.models.holding import Holding
from app.models.stock import Stock
from app.models.portfolio_history import PortfolioHistory
from app.models.rebalancing_history import RebalancingHistory


class ImportExportService:
    """Service for handling import/export of portfolios, strategies, and holdings"""
    
    # ===== TEMPLATE GENERATION =====
    
    def generate_portfolio_template(self) -> BytesIO:
        """Generate CSV template for portfolio import"""
        output = BytesIO()
        output.write(b"Portfolio Name,Stock Symbols\n")
        output.write(b"My Portfolio,\"AAPL,MSFT,GOOGL\"\n")
        output.write(b"Tech Stocks,\"META,NVDA,AMD,INTC\"\n")
        output.seek(0)
        return output
    
    def generate_strategy_template(self) -> BytesIO:
        """Generate Excel template for strategy import with holdings"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Holdings"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = ["Stock Symbol", "Quantity", "Purchase Price", "Purchase Date", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Example data
        example_data = [
            ["AAPL", 10, 150.50, "2024-01-15", "Tech stock"],
            ["MSFT", 5, 380.25, "2024-02-01", "Software giant"],
            ["GOOGL", 8, 140.75, "2024-02-15", "Search engine"],
        ]
        
        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # ===== PORTFOLIO EXPORT =====
    
    async def create_portfolio_excel(
        self, 
        db: AsyncSession, 
        portfolio: Portfolio,
        stocks: List[Stock]
    ) -> BytesIO:
        """Generate Excel export for a single portfolio"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Portfolio Details"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(size=14, bold=True)
        
        # Portfolio info
        ws.cell(row=1, column=1, value="Portfolio Name:").font = title_font
        ws.cell(row=1, column=2, value=portfolio.name).font = title_font
        ws.cell(row=2, column=1, value="Created Date:")
        ws.cell(row=2, column=2, value=portfolio.created_at.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=3, column=1, value="Stock Count:")
        ws.cell(row=3, column=2, value=len(stocks))
        
        # Stock table headers
        headers = ["Symbol", "Name", "Sector", "Current Price", "Change %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Stock data
        for row_idx, stock in enumerate(stocks, 6):
            ws.cell(row=row_idx, column=1, value=stock.symbol)
            ws.cell(row=row_idx, column=2, value=stock.name)
            ws.cell(row=row_idx, column=3, value=stock.sector or "N/A")
            ws.cell(row=row_idx, column=4, value=stock.current_price)
            ws.cell(row=row_idx, column=5, value=stock.change_percent or 0)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    async def create_all_portfolios_excel(
        self, 
        db: AsyncSession, 
        portfolios: List[Portfolio],
        all_stocks: Dict[int, Stock]
    ) -> BytesIO:
        """Generate Excel export for all user portfolios"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for portfolio in portfolios:
            # Create sheet for each portfolio
            ws = wb.create_sheet(title=portfolio.name[:31])  # Excel sheet name limit
            
            # Portfolio info
            ws.cell(row=1, column=1, value="Portfolio:").font = Font(bold=True)
            ws.cell(row=1, column=2, value=portfolio.name)
            ws.cell(row=2, column=1, value="Created:")
            ws.cell(row=2, column=2, value=portfolio.created_at.strftime("%Y-%m-%d"))
            
            # Headers
            headers = ["Symbol", "Name", "Sector", "Current Price", "Change %"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Stock data
            stocks = [all_stocks[sid] for sid in portfolio.stock_ids if sid in all_stocks]
            for row_idx, stock in enumerate(stocks, 5):
                ws.cell(row=row_idx, column=1, value=stock.symbol)
                ws.cell(row=row_idx, column=2, value=stock.name)
                ws.cell(row=row_idx, column=3, value=stock.sector or "N/A")
                ws.cell(row=row_idx, column=4, value=stock.current_price)
                ws.cell(row=row_idx, column=5, value=stock.change_percent or 0)
            
            # Adjust widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    async def create_portfolio_history_excel(
        self, 
        db: AsyncSession,
        portfolio: Portfolio,
        history: List[PortfolioHistory]
    ) -> BytesIO:
        """Generate Excel export for portfolio history"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Portfolio History"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Title
        ws.cell(row=1, column=1, value=f"History: {portfolio.name}").font = Font(size=14, bold=True)
        
        # Headers
        headers = ["Date", "Action", "Description", "Changes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # History data
        for row_idx, entry in enumerate(history, 4):
            ws.cell(row=row_idx, column=1, value=entry.created_at.strftime("%Y-%m-%d %H:%M"))
            ws.cell(row=row_idx, column=2, value=entry.action)
            ws.cell(row=row_idx, column=3, value=entry.description)
            ws.cell(row=row_idx, column=4, value=str(entry.changes))
        
        # Adjust widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 50
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # ===== STRATEGY EXPORT =====
    
    async def create_strategy_excel(
        self,
        db: AsyncSession,
        strategy: Strategy,
        holdings: List[Holding],
        stocks: Dict[int, Stock],
        portfolios: Dict[int, Portfolio]
    ) -> BytesIO:
        """Generate Excel export for strategy with multiple sheets"""
        wb = Workbook()
        wb.remove(wb.active)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Sheet 1: Strategy Info
        ws_info = wb.create_sheet("Strategy Info")
        ws_info.cell(row=1, column=1, value="Strategy Name:").font = Font(bold=True)
        ws_info.cell(row=1, column=2, value=strategy.name)
        ws_info.cell(row=2, column=1, value="Total Funds:").font = Font(bold=True)
        ws_info.cell(row=2, column=2, value=strategy.total_funds)
        ws_info.cell(row=3, column=1, value="Remaining Cash:").font = Font(bold=True)
        ws_info.cell(row=3, column=2, value=strategy.remaining_cash)
        ws_info.cell(row=4, column=1, value="Created Date:").font = Font(bold=True)
        ws_info.cell(row=4, column=2, value=strategy.created_at.strftime("%Y-%m-%d %H:%M"))
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 25
        
        # Sheet 2: Portfolio Allocations
        ws_alloc = wb.create_sheet("Portfolio Allocations")
        alloc_headers = ["Portfolio Name", "Percentage", "Stock Symbols"]
        for col, header in enumerate(alloc_headers, 1):
            cell = ws_alloc.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, alloc in enumerate(strategy.portfolio_allocations, 2):
            portfolio = portfolios.get(alloc['portfolio_id'])
            ws_alloc.cell(row=row_idx, column=1, value=portfolio.name if portfolio else f"Portfolio {alloc['portfolio_id']}")
            ws_alloc.cell(row=row_idx, column=2, value=alloc['percentage'])
            
            # Format stock symbols (without percentages)
            stock_allocs = alloc.get('stock_allocations', {})
            symbols = ", ".join([stocks.get(int(sid), Stock()).symbol or str(sid) 
                                for sid in stock_allocs.keys()])
            ws_alloc.cell(row=row_idx, column=3, value=symbols)
        
        ws_alloc.column_dimensions['A'].width = 25
        ws_alloc.column_dimensions['B'].width = 15
        ws_alloc.column_dimensions['C'].width = 50
        
        # Sheet 3: Holdings
        ws_hold = wb.create_sheet("Holdings")
        hold_headers = ["Symbol", "Name", "Allocation %", "Sector", "Quantity", "Avg Price", "Current Value", "Purchase Date", "Notes"]
        for col, header in enumerate(hold_headers, 1):
            cell = ws_hold.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Create lookup for stock allocation percentages
        stock_allocation_map = {}
        for alloc in strategy.portfolio_allocations:
            stock_allocs = alloc.get('stock_allocations', {})
            for stock_id_str, percentage in stock_allocs.items():
                stock_allocation_map[int(stock_id_str)] = percentage
        
        for row_idx, holding in enumerate(holdings, 2):
            stock = stocks.get(holding.stock_id)
            if stock:
                ws_hold.cell(row=row_idx, column=1, value=stock.symbol)
                ws_hold.cell(row=row_idx, column=2, value=stock.name)
                ws_hold.cell(row=row_idx, column=3, value=stock_allocation_map.get(holding.stock_id, 0))
                ws_hold.cell(row=row_idx, column=4, value=stock.sector or "N/A")
            ws_hold.cell(row=row_idx, column=5, value=holding.quantity)
            ws_hold.cell(row=row_idx, column=6, value=holding.average_price)
            ws_hold.cell(row=row_idx, column=7, value=holding.current_value)
            ws_hold.cell(row=row_idx, column=8, value=holding.purchase_date.strftime("%Y-%m-%d") if holding.purchase_date else "")
            ws_hold.cell(row=row_idx, column=9, value=holding.notes or "")
        
        ws_hold.column_dimensions['A'].width = 12
        ws_hold.column_dimensions['B'].width = 25
        ws_hold.column_dimensions['C'].width = 12
        ws_hold.column_dimensions['D'].width = 20
        ws_hold.column_dimensions['E'].width = 12
        ws_hold.column_dimensions['F'].width = 12
        ws_hold.column_dimensions['G'].width = 15
        ws_hold.column_dimensions['H'].width = 15
        ws_hold.column_dimensions['I'].width = 30
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    async def create_strategy_history_excel(
        self,
        db: AsyncSession,
        strategy: Strategy,
        history: List[RebalancingHistory],
        stocks: Dict[int, Stock]
    ) -> BytesIO:
        """Generate Excel export for strategy rebalancing history"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rebalancing History"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Title
        ws.cell(row=1, column=1, value=f"Rebalancing History: {strategy.name}").font = Font(size=14, bold=True)
        
        # Headers
        headers = ["Date", "Executed", "Undone", "Actions"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # History data
        for row_idx, entry in enumerate(history, 4):
            ws.cell(row=row_idx, column=1, value=entry.created_at.strftime("%Y-%m-%d %H:%M"))
            ws.cell(row=row_idx, column=2, value="Yes" if entry.executed else "No")
            ws.cell(row=row_idx, column=3, value="Yes" if entry.undone else "No")
            
            # Format actions
            actions_str = ""
            if isinstance(entry.actions, list):
                for action in entry.actions:
                    stock_id = action.get('stock_id')
                    stock = stocks.get(stock_id)
                    symbol = stock.symbol if stock else f"Stock {stock_id}"
                    actions_str += f"{action.get('action', '')}: {symbol} x{action.get('quantity', 0)} @ ${action.get('price', 0)}; "
            ws.cell(row=row_idx, column=4, value=actions_str)
        
        # Adjust widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 60
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # ===== HOLDINGS EXPORT =====
    
    async def create_holdings_excel(
        self,
        db: AsyncSession,
        holdings: List[Holding],
        stocks: Dict[int, Stock],
        strategies: Dict[int, Strategy],
        portfolios: Dict[int, Portfolio]
    ) -> BytesIO:
        """Generate Excel export for all holdings"""
        wb = Workbook()
        ws = wb.active
        ws.title = "All Holdings"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = [
            "Symbol", "Name", "Sector", "Quantity", "Avg Price", "Current Price", 
            "Current Value", "Total P/L", "P/L %", "Strategy", "Portfolio", 
            "Purchase Date", "Notes", "Manual"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Holdings data
        for row_idx, holding in enumerate(holdings, 2):
            stock = stocks.get(holding.stock_id)
            if stock:
                ws.cell(row=row_idx, column=1, value=stock.symbol)
                ws.cell(row=row_idx, column=2, value=stock.name)
                ws.cell(row=row_idx, column=3, value=stock.sector or "N/A")
                ws.cell(row=row_idx, column=6, value=stock.current_price)
                
                # Calculate P/L
                total_pl = (stock.current_price - holding.average_price) * holding.quantity
                pl_percent = ((stock.current_price - holding.average_price) / holding.average_price * 100) if holding.average_price > 0 else 0
                
                ws.cell(row=row_idx, column=8, value=total_pl)
                ws.cell(row=row_idx, column=9, value=pl_percent)
            
            ws.cell(row=row_idx, column=4, value=holding.quantity)
            ws.cell(row=row_idx, column=5, value=holding.average_price)
            ws.cell(row=row_idx, column=7, value=holding.current_value)
            
            # Strategy and Portfolio names
            strategy = strategies.get(holding.strategy_id) if holding.strategy_id else None
            portfolio = portfolios.get(holding.portfolio_id) if holding.portfolio_id else None
            ws.cell(row=row_idx, column=10, value=strategy.name if strategy else "")
            ws.cell(row=row_idx, column=11, value=portfolio.name if portfolio else "")
            
            ws.cell(row=row_idx, column=12, value=holding.purchase_date.strftime("%Y-%m-%d") if holding.purchase_date else "")
            ws.cell(row=row_idx, column=13, value=holding.notes or "")
            ws.cell(row=row_idx, column=14, value="Yes" if holding.is_manual else "No")
        
        # Adjust column widths
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['B'].width = 30  # Name
        ws.column_dimensions['C'].width = 20  # Sector
        ws.column_dimensions['M'].width = 30  # Notes
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # ===== COMPARISON EXPORT =====
    
    async def create_comparison_excel(
        self,
        db: AsyncSession,
        comparison_data: List[Dict[str, Any]]
    ) -> BytesIO:
        """Generate Excel export for portfolio comparison"""
        wb = Workbook()
        wb.remove(wb.active)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Comparison Summary")
        summary_headers = ["Portfolio Name", "Current Value", "Initial Value", "Change", "Change %", "Stock Count"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, data in enumerate(comparison_data, 2):
            ws_summary.cell(row=row_idx, column=1, value=data.get('portfolio_name', ''))
            ws_summary.cell(row=row_idx, column=2, value=data.get('current_value', 0))
            ws_summary.cell(row=row_idx, column=3, value=data.get('initial_value', 0))
            ws_summary.cell(row=row_idx, column=4, value=data.get('change', 0))
            ws_summary.cell(row=row_idx, column=5, value=data.get('change_percent', 0))
            ws_summary.cell(row=row_idx, column=6, value=data.get('stock_count', 0))
        
        ws_summary.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws_summary.column_dimensions[col].width = 15
        
        # Sheet 2+: Individual portfolio time series
        for data in comparison_data:
            portfolio_name = data.get('portfolio_name', 'Portfolio')
            time_series = data.get('time_series', [])
            
            if time_series:
                ws = wb.create_sheet(portfolio_name[:31])
                
                # Headers
                ts_headers = ["Date", "Value"]
                for col, header in enumerate(ts_headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Data
                for row_idx, point in enumerate(time_series, 2):
                    ws.cell(row=row_idx, column=1, value=point.get('date', ''))
                    ws.cell(row=row_idx, column=2, value=point.get('value', 0))
                
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # ===== IMPORT PARSING =====
    
    async def parse_portfolio_csv(self, file_content: bytes) -> List[Dict[str, Any]]:
        """Parse portfolio CSV file"""
        content = file_content.decode('utf-8')
        csv_reader = csv.DictReader(content.splitlines())
        
        portfolios = []
        for row in csv_reader:
            portfolio_name = row.get('Portfolio Name', '').strip()
            stock_symbols = row.get('Stock Symbols', '').strip()
            
            if portfolio_name and stock_symbols:
                # Split stock symbols by comma
                symbols = [s.strip() for s in stock_symbols.split(',') if s.strip()]
                portfolios.append({
                    'name': portfolio_name,
                    'symbols': symbols
                })
        
        return portfolios
    
    async def parse_strategy_excel(
        self, 
        file_content: bytes,
        db: AsyncSession
    ) -> Dict[str, Any]:
        """Parse strategy Excel file with holdings data"""
        wb = load_workbook(BytesIO(file_content))
        ws = wb.active
        
        holdings = []
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            symbol = str(row[0]).strip().upper()
            quantity = int(row[1]) if row[1] else 0
            purchase_price = float(row[2]) if row[2] else 0.0
            purchase_date_str = str(row[3]) if row[3] else None
            notes = str(row[4]) if row[4] else ""
            
            # Parse date
            purchase_date = None
            if purchase_date_str:
                try:
                    if isinstance(row[3], datetime):
                        purchase_date = row[3]
                    else:
                        purchase_date = datetime.strptime(purchase_date_str, "%Y-%m-%d")
                except ValueError:
                    pass  # Invalid date format, leave as None
            
            holdings.append({
                'symbol': symbol,
                'quantity': quantity,
                'purchase_price': purchase_price,
                'purchase_date': purchase_date,
                'notes': notes
            })
        
        return {'holdings': holdings}


# Singleton instance
import_export_service = ImportExportService()

